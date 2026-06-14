#!/usr/bin/env python3
"""
Amazon ASIN 商品情報取得ツール — Python版（PA-API v5）
======================================================
動作環境: Mac（python3） / Windows（python）
出力形式: CSV または Excel（.xlsx）

使い方:
  1. 依存ライブラリをインストール
       pip install requests openpyxl  （または pip3）
  2. 下記 CONFIG に PA-API キーを設定（または .env ファイル）
  3. ASINをファイル（asin_list.txt）またはコマンドライン引数で渡す
  4. 実行
       python3 asin_fetcher.py asin_list.txt --output result.xlsx

注意:
  - 実際の API 呼び出しは PA-API キー設定後に有効になります
  - キー未設定時はダミーデータを返します（テスト用）
  - .env ファイルや環境変数での設定を強く推奨します（コードへの直書き禁止）
"""

import os
import sys
import csv
import hmac
import json
import time
import hashlib
import datetime
import argparse
import pathlib
from typing import Optional

# ----------- 依存ライブラリ（pip install requests openpyxl） -----------
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("[警告] requests がインストールされていません。`pip install requests` を実行してください。")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# 設定（環境変数 or .env ファイルから読み込む）
# ============================================================
def load_env():
    """プロジェクトルートの .env ファイルを読み込む（python-dotenv なしの簡易版）"""
    env_path = pathlib.Path(__file__).parent / ".env"
    if env_path.exists():
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, val = line.partition("=")
                    os.environ.setdefault(key.strip(), val.strip())

load_env()

CONFIG = {
    "access_key":  os.environ.get("PA_ACCESS_KEY",  "YOUR_ACCESS_KEY_HERE"),
    "secret_key":  os.environ.get("PA_SECRET_KEY",  "YOUR_SECRET_KEY_HERE"),
    "partner_tag": os.environ.get("PA_PARTNER_TAG", "your-associate-tag-22"),
    "host":        "webservices.amazon.co.jp",
    "region":      "us-west-2",
    "batch_size":  10,         # PA-API 最大 10件/リクエスト
    "wait_sec":    1.1,        # レート制限対策: リクエスト間のウェイト
}

PA_RESOURCES = [
    "ItemInfo.Title",
    "ItemInfo.Features",
    "ItemInfo.ProductInfo",
    "ItemInfo.ByLineInfo",
    "Offers.Listings.Price",
    "Offers.Listings.Availability.Type",
    "Images.Primary.Large",
    "VariationSummary.VariationDimension",
]

DUMMY_DATA = {
    "B001AABBCC": {"title": "ステンレス真空断熱ボトル 500ml シルバー",       "price": 2980, "brand": "サンプルブランド", "image": "https://example.com/img1.jpg", "weight": "280g", "features": "保温保冷 / 広口 / 食洗機対応", "variations": "カラー, サイズ", "availability": "在庫あり"},
    "B002CCDDEE": {"title": "USB-C ハブ 7-in-1 マルチポートアダプター",       "price": 3480, "brand": "TechSample",      "image": "https://example.com/img2.jpg", "weight": "68g",  "features": "HDMI 4K / PD 100W / USB3.0×3", "variations": "カラー", "availability": "在庫あり"},
    "B003EEFFGG": {"title": "ワイヤレスイヤホン Bluetooth 5.3 ノイズキャンセリング", "price": 6800, "brand": "AudioSample", "image": "https://example.com/img3.jpg", "weight": "55g",  "features": "ANC搭載 / 最大30時間 / IPX5", "variations": "カラー", "availability": "在庫あり"},
    "B004HHIIJJ": {"title": "ポータブル充電器 20000mAh PD対応",                "price": 4200, "brand": "PowerSample",    "image": "https://example.com/img4.jpg", "weight": "430g", "features": "PD 65W / ノートPC対応 / 大容量", "variations": "", "availability": "残りわずか"},
}


# ============================================================
# AWS Signature Version 4
# ============================================================
def _sha256(data: str) -> str:
    return hashlib.sha256(data.encode("utf-8")).hexdigest()


def _hmac_sha256(key: bytes, data: str) -> bytes:
    return hmac.new(key, data.encode("utf-8"), hashlib.sha256).digest()


def _sign(key: bytes, msg: str) -> bytes:
    return _hmac_sha256(key, msg)


def _get_signing_key(secret_key: str, date_stamp: str, region: str, service: str) -> bytes:
    k_date    = _hmac_sha256(f"AWS4{secret_key}".encode("utf-8"), date_stamp)
    k_region  = _hmac_sha256(k_date, region)
    k_service = _hmac_sha256(k_region, service)
    k_signing = _hmac_sha256(k_service, "aws4_request")
    return k_signing


def build_auth_header(body: str, amz_date: str, date_stamp: str) -> dict:
    """PA-API v5 用の Authorization ヘッダーを生成する"""
    service     = "ProductAdvertisingAPI"
    target      = "com.amazon.paapi5.v1.ProductAdvertisingAPIv1.GetItems"
    payload_hash = _sha256(body)

    # 正規ヘッダー（アルファベット順）
    headers = {
        "content-encoding": "amz-1.0",
        "content-type":     "application/json; charset=utf-8",
        "host":             CONFIG["host"],
        "x-amz-date":      amz_date,
        "x-amz-target":    target,
    }
    sorted_keys       = sorted(headers.keys())
    canonical_headers = "".join(f"{k}:{headers[k]}\n" for k in sorted_keys)
    signed_headers    = ";".join(sorted_keys)

    canonical_request = "\n".join([
        "POST",
        "/paapi5/getitems",
        "",
        canonical_headers,
        signed_headers,
        payload_hash,
    ])

    credential_scope = f"{date_stamp}/{CONFIG['region']}/{service}/aws4_request"
    string_to_sign   = "\n".join([
        "AWS4-HMAC-SHA256",
        amz_date,
        credential_scope,
        _sha256(canonical_request),
    ])

    signing_key = _get_signing_key(CONFIG["secret_key"], date_stamp, CONFIG["region"], service)
    signature   = _hmac_sha256(signing_key, string_to_sign).hex()

    authorization = (
        f"AWS4-HMAC-SHA256 "
        f"Credential={CONFIG['access_key']}/{credential_scope}, "
        f"SignedHeaders={signed_headers}, "
        f"Signature={signature}"
    )

    return {**headers, "Authorization": authorization}


# ============================================================
# PA-API 呼び出し
# ============================================================
def call_paapi(asins: list[str]) -> dict:
    """ASIN リストを PA-API に送信し、商品情報を dict で返す"""
    if not HAS_REQUESTS:
        raise RuntimeError("requests ライブラリが必要です: pip install requests")

    # APIキー未設定 → ダミーを返す
    if CONFIG["access_key"] == "YOUR_ACCESS_KEY_HERE":
        print("  [INFO] PA-APIキー未設定 → ダミーデータを返します")
        return {asin: _build_dummy_item(asin) for asin in asins}

    payload = {
        "ItemIds":     asins,
        "Resources":   PA_RESOURCES,
        "PartnerTag":  CONFIG["partner_tag"],
        "PartnerType": "Associates",
        "Marketplace": "www.amazon.co.jp",
    }
    body     = json.dumps(payload)
    now      = datetime.datetime.utcnow()
    amz_date  = now.strftime("%Y%m%dT%H%M%SZ")
    date_stamp = now.strftime("%Y%m%d")

    auth_headers = build_auth_header(body, amz_date, date_stamp)

    resp = requests.post(
        f"https://{CONFIG['host']}/paapi5/getitems",
        headers=auth_headers,
        data=body.encode("utf-8"),
        timeout=15,
    )

    if resp.status_code != 200:
        raise RuntimeError(f"PA-API エラー [{resp.status_code}]: {resp.text[:200]}")

    items_dict = {}
    for item in resp.json().get("ItemsResult", {}).get("Items", []):
        items_dict[item["ASIN"]] = item
    return items_dict


def _build_dummy_item(asin: str) -> dict:
    d = DUMMY_DATA.get(asin, {
        "title": f"[ダミー] 商品 {asin}",
        "price": 1000,
        "brand": "テストブランド",
        "image": "",
        "weight": "不明",
        "features": "テストデータ",
        "variations": "",
        "availability": "不明",
    })
    return {"_dummy": True, **d}


def parse_item(asin: str, item: dict) -> dict:
    """PA-API レスポンスまたはダミーを共通フォーマットに変換"""
    if item.get("_dummy"):
        return {
            "asin":         asin,
            "title":        item.get("title", ""),
            "price":        item.get("price", ""),
            "currency":     "JPY",
            "brand":        item.get("brand", ""),
            "image_url":    item.get("image", ""),
            "features":     item.get("features", ""),
            "variations":   item.get("variations", ""),
            "weight":       item.get("weight", ""),
            "availability": item.get("availability", ""),
            "detail_url":   f"https://www.amazon.co.jp/dp/{asin}",
            "status":       "成功（ダミー）",
            "fetched_at":   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    info    = item.get("ItemInfo", {})
    offers  = item.get("Offers", {}).get("Listings", [{}])
    listing = offers[0] if offers else {}
    images  = item.get("Images", {}).get("Primary", {})
    prod    = info.get("ProductInfo", {})
    wt      = prod.get("ItemDimensions", {}).get("Weight", {})

    return {
        "asin":         asin,
        "title":        info.get("Title", {}).get("DisplayValue", "取得不可"),
        "price":        listing.get("Price", {}).get("Amount", ""),
        "currency":     listing.get("Price", {}).get("Currency", "JPY"),
        "brand":        info.get("ByLineInfo", {}).get("Brand", {}).get("DisplayValue", "取得不可"),
        "image_url":    images.get("Large", {}).get("URL", "取得不可"),
        "features":     " / ".join((info.get("Features", {}).get("DisplayValues", []) or [])[:3]) or "取得不可",
        "variations":   " / ".join(item.get("VariationSummary", {}).get("VariationDimension", [])) or "なし",
        "weight":       f"{wt.get('DisplayValue', '')} {wt.get('Unit', '')}" if wt else "取得不可",
        "availability": listing.get("Availability", {}).get("Type", "不明"),
        "detail_url":   f"https://www.amazon.co.jp/dp/{asin}",
        "status":       "成功",
        "fetched_at":   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ============================================================
# 出力
# ============================================================
COLUMNS = ["asin", "title", "price", "currency", "brand", "image_url",
           "features", "variations", "weight", "availability", "detail_url",
           "status", "fetched_at"]
COL_LABELS = ["ASIN", "商品名", "価格(円)", "通貨", "ブランド", "画像URL",
              "商品説明", "バリエーション", "重量", "在庫状況", "商品ページURL",
              "ステータス", "取得日時"]


def write_csv(rows: list[dict], path: pathlib.Path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writerow(dict(zip(COLUMNS, COL_LABELS)))
        writer.writerows(rows)
    print(f"[完了] CSV 出力: {path}")


def write_excel(rows: list[dict], path: pathlib.Path):
    if not HAS_OPENPYXL:
        print("[警告] openpyxl がありません。CSV にフォールバックします。")
        write_csv(rows, path.with_suffix(".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "取得結果"

    # ヘッダー行
    header_fill = PatternFill("solid", fgColor="131921")
    for col_idx, label in enumerate(COL_LABELS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    # 列幅を自動調整（簡易版）
    col_widths = [12, 40, 10, 8, 20, 50, 40, 20, 10, 12, 40, 10, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[完了] Excel 出力: {path}")


# ============================================================
# メイン
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Amazon ASIN 商品情報取得ツール（PA-API v5）"
    )
    parser.add_argument("asin_file", nargs="?",
                        help="ASINが1行1件で書かれたテキストファイル（省略時はサンプルASINを使用）")
    parser.add_argument("--output", "-o", default="asin_result.xlsx",
                        help="出力ファイル名（.xlsx または .csv、デフォルト: asin_result.xlsx）")
    args = parser.parse_args()

    # ASINリスト読み込み
    if args.asin_file:
        asin_path = pathlib.Path(args.asin_file)
        if not asin_path.exists():
            print(f"[エラー] ファイルが見つかりません: {asin_path}")
            sys.exit(1)
        text = asin_path.read_text(encoding="utf-8")
    else:
        print("[INFO] ASINファイルが指定されていないため、サンプルASINで実行します")
        text = "B001AABBCC\nB002CCDDEE\nB003EEFFGG\nB004HHIIJJ\nB005KKLLMM"

    asins = [
        line.strip().upper()
        for line in text.splitlines()
        if line.strip() and len(line.strip()) == 10
    ]
    if not asins:
        print("[エラー] 有効なASIN（10桁英数字）が見つかりませんでした")
        sys.exit(1)

    print(f"[開始] {len(asins)} 件のASINを処理します")

    output_path = pathlib.Path(args.output)
    results = []
    errors  = []

    for i in range(0, len(asins), CONFIG["batch_size"]):
        batch = asins[i:i + CONFIG["batch_size"]]
        print(f"  バッチ {i+1}〜{i+len(batch)} / {len(asins)} を取得中...")
        try:
            items = call_paapi(batch)
            for asin in batch:
                if asin in items:
                    results.append(parse_item(asin, items[asin]))
                else:
                    results.append({
                        "asin": asin, "title": "商品情報なし（削除・非対応の可能性）",
                        "status": "エラー", "fetched_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                    errors.append(asin)
        except Exception as e:
            print(f"  [エラー] {e}")
            for asin in batch:
                results.append({
                    "asin": asin, "title": "APIエラー",
                    "status": f"エラー: {e}", "fetched_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
                errors.append(asin)

        # レート制限対策
        if i + CONFIG["batch_size"] < len(asins):
            time.sleep(CONFIG["wait_sec"])

    # 出力
    if output_path.suffix.lower() in (".xlsx", ".xls"):
        write_excel(results, output_path)
    else:
        write_csv(results, output_path)

    success = len(results) - len(errors)
    print(f"[サマリー] 成功: {success}件 / エラー: {len(errors)}件")
    if errors:
        print(f"  エラーASIN: {', '.join(errors)}")


if __name__ == "__main__":
    main()
